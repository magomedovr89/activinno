import openpyxl
wookbook = openpyxl.load_workbook("18.xlsx")
worksheet = wookbook.active
data = []
for i in range(0, worksheet.max_row):
    temporary = []
    for col in worksheet.iter_cols(1, worksheet.max_column):
        temporary.append(col[i].value)
    data.append(temporary)
print(data[0][0])
print(len(data))
